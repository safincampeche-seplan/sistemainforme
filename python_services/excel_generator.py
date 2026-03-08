import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image
from io import BytesIO
import os
from models import ExcelExportRequest, ConsolidatedExcelExportRequest

def generate_styled_excel(data: ExcelExportRequest) -> BytesIO:
    # 1. Crear DataFrame
    df = pd.DataFrame(data.rows)
    
    # Asegurar orden de columnas segun headers si es posible
    # (Si los keys en rows coinciden con IDs de headers)
    
    output = BytesIO()
    
    # 2. Usar ExcelWriter con engine openpyxl
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, startrow=5, sheet_name='Anexo Estadístico')
        
        workbook = writer.book
        worksheet = writer.sheets['Anexo Estadístico']
        
        # --- ESTILIZACIÓN ---
        
        # Logo (si existe)
        logo_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "../public/images/logo_semaig.png"))
        if os.path.exists(logo_path):
            try:
                img = Image(logo_path)
                img.height = 60
                img.width = 150
                worksheet.add_image(img, 'A1')
            except: pass

        # Título Principal
        title_cell = worksheet.cell(row=2, column=3)
        title_cell.value = data.report_title
        title_cell.font = Font(size=18, bold=True, color="1E293B")
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Subtítulo (Entidad)
        subtitle_cell = worksheet.cell(row=3, column=3)
        subtitle_cell.value = f"Entidad: {data.entity_name}"
        subtitle_cell.font = Font(size=12, bold=True, color="475569")
        
        # Metadatos del Informe
        worksheet.cell(row=4, column=3).value = f"Fecha de Generación: {pd.Timestamp.now().strftime('%d/%m/%Y %H:%M')}"
        worksheet.cell(row=4, column=3).font = Font(size=10, italic=True)

        # Estilo de Cabeceras de Tabla
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Sobrescribir cabeceras con nombres legibles si se proporcionan
        for col_num, header_name in enumerate(data.headers, 1):
            cell = worksheet.cell(row=6, column=col_num)
            cell.value = header_name
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        # Ajustar ancho de columnas y bordes de datos
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.row >= 6:
                    cell.border = thin_border
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except: pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column].width = min(adjusted_width, 50)

    output.seek(0)
    return output

def generate_consolidated_styled_excel(data: ConsolidatedExcelExportRequest) -> BytesIO:
    output = BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for item in data.items:
            # Limitar el nombre de la hoja a 31 caracteres maximo y remover caracteres inválidos para Excel
            import re
            sheet_title = re.sub(r'[\\*?:/\[\]]', '', item.entity_name)[:31].strip()
            if not sheet_title:
                sheet_title = "Sheet"
                
            df = pd.DataFrame(item.rows)
            df.to_excel(writer, index=False, startrow=5, sheet_name=sheet_title)
            
            workbook = writer.book
            # Acceder a la ultima hoja creada en lugar de por nombre exacto, 
            # ya que pandas puede alterar sheet_title si hay colisiones.
            worksheet = workbook.worksheets[-1]
            
            # --- ESTILIZACIÓN ---
            logo_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "../public/images/logo_semaig.png"))
            if os.path.exists(logo_path):
                try:
                    img = Image(logo_path)
                    img.height = 60
                    img.width = 150
                    worksheet.add_image(img, 'A1')
                except: pass

            # Título Principal
            title_cell = worksheet.cell(row=2, column=3)
            title_cell.value = data.report_title
            title_cell.font = Font(size=18, bold=True, color="1E293B")
            title_cell.alignment = Alignment(horizontal="left", vertical="center")
            
            # Subtítulo (Entidad)
            subtitle_cell = worksheet.cell(row=3, column=3)
            subtitle_cell.value = f"Entidad: {item.entity_name}"
            subtitle_cell.font = Font(size=12, bold=True, color="475569")
            
            # Metadatos del Informe
            worksheet.cell(row=4, column=3).value = f"Fecha de Generación: {pd.Timestamp.now().strftime('%d/%m/%Y %H:%M')}"
            worksheet.cell(row=4, column=3).font = Font(size=10, italic=True)

            # Estilo de Cabeceras
            header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            
            for col_num, header_name in enumerate(item.headers, 1):
                cell = worksheet.cell(row=6, column=col_num)
                cell.value = header_name
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin_border

            # Ajustar ancho de columnas y bordes
            for col in worksheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    if cell.row >= 6:
                        cell.border = thin_border
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except: pass
                worksheet.column_dimensions[column].width = min((max_length + 2), 50)

    output.seek(0)
    return output
